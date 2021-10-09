
#connect to PowerFactory
import powerfactory as pf

app = pf.GetApplication()

import random

for dataset_n in range(102):
  
  random.seed(dataset_n+1)
  
  #Library for connecting python and excel API
  import openpyxl as xl
  from openpyxl.chart import LineChart, Reference
  
  app.ClearOutputWindow()
  
  #get active project
  prj = app.GetActiveProject()
  filename = prj.GetAttribute("loc_name")
  
  pq_bus = ['bus 01','bus 02','bus 03','bus 04','bus 05','bus 06','bus 07','bus 08','bus 09','bus 10','bus 11','bus 12','bus 13','bus 14']
  
  P_list = [2.4,2.4,2.4,2.4,2.4,2.4,2.4,2.4,2.4,2.4,2.4,2.4,2.4,2.4]
  Q_list = [0.1,0.1,0.1,0.1,0.1,0.1,0.1,0.1,0.1,0.1,0.1,0.1,0.1,0.1]
  
  #get all loads
  loads = app.GetCalcRelevantObjects('*.ElmLod')
  load_dict = {}
  for i,load in enumerate(loads):
    load_dict[load.loc_name] = load
    load_dict[load.loc_name].SetAttribute("plini", P_list[i])
    load_dict[load.loc_name].SetAttribute("qlini", Q_list[i])
  
  #get all buses
  buses = app.GetCalcRelevantObjects('*.ElmTerm')
  bus_dict = {}
  for bus in buses:
    if bus.loc_name in pq_bus:
      bus_dict[bus.loc_name] = bus
  
  #retrieve load-flow object
  ldf = app.GetFromStudyCase("ComLdf")
  
  wb = xl.Workbook()
  sheet = wb.active
  sheet.title = 'PowerFlowData'
  
  sheet['A1'] = 'PF Dataset_{:d}'.format(dataset_n+1)
  
  for n,bus_key in enumerate(bus_dict.keys()):
    
    sheet.cell(row = 1, column = n*4+2).value = ("P_%s (PQ)" % (n+1))
    sheet.cell(row = 1, column = n*4+3).value = ("Q_%s (PQ)" % (n+1))
    sheet.cell(row = 1, column = n*4+4).value = ("V_%s (PQ)" % (n+1))
    sheet.cell(row = 1, column = n*4+5).value = ("d_%s (PQ)" % (n+1))
      
  for i in range(2000):
    app.PrintPlain("ITERATION %s" % (i+1))
    
    sheet.cell(row = i+2, column = 1).value = 'Data ' + str(i+1)
    
    for n,load_key in enumerate(load_dict.keys()):
      load_dict[load_key].SetAttribute("plini", P_list[n] + P_list[n]*random.uniform(-0.5, 0.5))
      load_dict[load_key].SetAttribute("qlini", Q_list[n] + Q_list[n]*random.uniform(-0.5, 0.5))
    
    #force balanced load flow  
    ldf.iopt_net = 0
    #execute load flow
    ldf.Execute()
    
    #for n, bus in enumerate(buses):
    for n,bus_key in enumerate(bus_dict.keys()):

      P_load = bus_dict[bus_key].GetAttribute("m:Pload")
      Q_load = bus_dict[bus_key].GetAttribute("m:Qload")
      V_load = bus_dict[bus_key].GetAttribute("m:Ul")
      d_load = bus_dict[bus_key].GetAttribute("m:phiu")
      
      sheet.cell(row = i+2, column = n*4+2).value = '{0:.3f}'.format(P_load)
      sheet.cell(row = i+2, column = n*4+3).value = '{0:.3f}'.format(Q_load)
      sheet.cell(row = i+2, column = n*4+4).value = '{0:.3f}'.format(V_load)
      sheet.cell(row = i+2, column = n*4+5).value = '{0:.3f}'.format(d_load)
        
  #save the excel file
  wb.save('C:\\Users\\ASUS\\Desktop\\%s_%s.xlsx' % (filename, dataset_n+1))